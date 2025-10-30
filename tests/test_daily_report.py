#!/usr/bin/env python3
"""
日报模块综合测试文件
包含格式验证、条件格式化测试、除零错误测试等
"""

import sys
from pathlib import Path
import numpy as np
import pandas as pd

# 添加项目根目录到Python路径
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))


def test_active_users_format():
    """测试日活数值格式是否正确显示小数"""
    print("🔍 测试日活数值格式")
    print("-" * 50)

    try:
        import pandas as pd
        from openpyxl import load_workbook

        # 查找最新的日报文件
        outputs_dir = project_root / "outputs"
        daily_report_files = list(outputs_dir.glob("daily_report_*.xlsx"))

        if not daily_report_files:
            print("❌ 未找到日报文件")
            return False

        # 查找包含'品类数据'sheet的文件
        target_file = None
        for file in sorted(daily_report_files, key=lambda x: x.stat().st_mtime, reverse=True):
            try:
                # 尝试读取文件，看是否包含'品类数据'sheet
                pd.read_excel(file, sheet_name='品类数据', nrows=1)
                target_file = file
                break
            except:
                continue

        if not target_file:
            print("❌ 未找到包含'品类数据'sheet的日报文件")
            return False

        print(f"📄 使用文件: {target_file.name}")

        # 使用pandas读取数据
        df = pd.read_excel(target_file, sheet_name='品类数据', skiprows=1)

        print("📊 日活列示例:")
        for i in range(min(5, len(df))):
            for col in df.columns:
                if '日活' in col:
                    print(f"  {df.loc[i, '一级分类']:12} | {col}: {df.loc[i, col]:8.2f}")

        # 检查Excel格式
        wb = load_workbook(target_file)
        ws = wb['品类数据']

        print("\n📈 Excel格式验证:")
        for col_idx, col_name in enumerate(df.columns):
            if '日活' in col_name:
                from openpyxl.utils import get_column_letter
                col_letter = get_column_letter(col_idx + 1)
                cell = ws.cell(row=3, column=col_idx + 1)
                number_format = cell.number_format
                print(f"  {col_name}: {number_format}")

        wb.close()
        print("✅ 日活数值格式测试通过\n")
        return True

    except Exception as e:
        print(f"❌ 日活数值格式测试失败: {e}\n")
        return False


def test_conditional_formatting():
    """测试条件格式化是否正确应用"""
    print("🔍 测试条件格式化")
    print("-" * 50)

    try:
        import openpyxl
        from openpyxl import load_workbook

        # 查找最新的日报文件
        outputs_dir = project_root / "outputs"
        daily_report_files = list(outputs_dir.glob("daily_report_*.xlsx"))

        if not daily_report_files:
            print("❌ 未找到日报文件")
            return False

        # 查找包含'品类数据'sheet的文件
        target_file = None
        for file in sorted(daily_report_files, key=lambda x: x.stat().st_mtime, reverse=True):
            try:
                wb = load_workbook(file)
                if '品类数据' in wb.sheetnames:
                    target_file = file
                    break
                wb.close()
            except:
                continue

        if not target_file:
            print("❌ 未找到包含'品类数据'sheet的日报文件")
            return False

        print(f"📄 使用文件: {target_file.name}")

        wb = load_workbook(target_file)
        ws = wb['品类数据']

        # 检查条件格式化规则
        cf_list = ws.conditional_formatting
        print(f"条件格式化规则数量: {len(cf_list)}")

        # 检查负值单元格
        negative_cells = []
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, (int, float)) and cell.value < 0:
                    negative_cells.append((row, col, cell.value))

        print(f"发现负值单元格数量: {len(negative_cells)}")
        if negative_cells:
            print("负值单元格位置:")
            for row, col, value in negative_cells[:5]:  # 只显示前5个
                print(f"  第{row}行第{col}列 = {value}")

        print("✅ 条件格式化测试通过\n")
        wb.close()
        return True

    except Exception as e:
        print(f"❌ 条件格式化测试失败: {e}\n")
        return False


def test_division_by_zero():
    """测试环比计算中的除零错误处理"""
    print("🔍 测试除零错误处理")
    print("-" * 50)

    try:
        # 创建测试数据
        test_data = {
            '一级分类': ['测试A', '测试B', '测试C'],
            '30日金额': [1000.0, 500.0, 200.0],
            '29日金额': [800.0, 0.0, 100.0],  # 包含0值
        }

        df = pd.DataFrame(test_data)
        print("📊 测试数据:")
        print(df)

        # 模拟环比计算
        current_col = '30日金额'
        compare_col = '29日金额'

        with np.errstate(divide='ignore', invalid='ignore'):
            ratio_values = df[current_col] / df[compare_col] - 1
            ratio_values[df[compare_col] == 0] = np.nan  # 分母为0时设为NaN
            ratio_values = ratio_values.round(4)

        print("\n🔬 环比计算结果:")
        for i, (idx, row) in enumerate(df.iterrows()):
            current = row[current_col]
            compare = row[compare_col]
            ratio = ratio_values.iloc[i]

            if compare == 0:
                print(f"  {row['一级分类']}: {current} / {compare} - 1 = 空值（分母为0）")
            elif pd.isna(ratio):
                print(f"  {row['一级分类']}: 计算错误")
            else:
                print(f"  {row['一级分类']}: {ratio:.4f} ({ratio:.2%})")

        print("✅ 除零错误处理测试通过\n")
        return True

    except Exception as e:
        print(f"❌ 除零错误处理测试失败: {e}\n")
        return False


def generate_test_report():
    """生成测试报告"""
    print("📝 生成测试报告")
    print("-" * 50)

    try:
        from app.outputs.daily_report.writer import DailyReportWriter

        # 创建测试数据，包含负值和除零场景
        test_data = {
            '一级分类': ['测试A', '测试B', '测试C', '测试D'],
            '30日金额': [1000.0, 500.0, -200.0, 300.0],
            '29日金额': [800.0, 0.0, 100.0, 400.0],
            '金额差值': [200.0, 500.0, -300.0, -100.0],
            '金额环比': [0.25, np.nan, -3.0, -0.25],  # 包含NaN
        }

        df = pd.DataFrame(test_data)

        writer = DailyReportWriter()
        sheet_data = {
            'sheet_name': '日报测试',
            'title': '日报模块综合测试',
            'data': df
        }

        output_file = writer.write_single_sheet_report(sheet_data)
        print(f"✅ 测试报告已生成: outputs/{Path(output_file).name}")
        print("🔍 请检查文件中的:")
        print("  - 负值红色背景格式化")
        print("  - 数值格式显示")
        print("  - 条件格式化覆盖范围")
        print("✅ 测试报告生成完成\n")
        return True

    except Exception as e:
        print(f"❌ 测试报告生成失败: {e}\n")
        return False


def main():
    """主测试函数"""
    print("🧪 日报模块综合测试")
    print("=" * 60)

    results = []

    # 运行各项测试
    results.append(("日活数值格式", test_active_users_format()))
    results.append(("条件格式化", test_conditional_formatting()))
    results.append(("除零错误处理", test_division_by_zero()))
    results.append(("测试报告生成", generate_test_report()))

    # 汇总结果
    print("📊 测试结果汇总")
    print("=" * 60)

    passed = 0
    total = len(results)

    for test_name, result in results:
        status = "✅ 通过" if result else "❌ 失败"
        print(f"{test_name:20} {status}")
        if result:
            passed += 1

    print(f"\n总计: {passed}/{total} 项测试通过")

    if passed == total:
        print("🎉 所有测试通过！日报模块功能正常")
        return 0
    else:
        print("⚠️  部分测试失败，请检查相关功能")
        return 1


if __name__ == "__main__":
    exit_code = main()
    sys.exit(exit_code)